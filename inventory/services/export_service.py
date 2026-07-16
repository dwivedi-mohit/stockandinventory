import csv
import os
from datetime import datetime
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, Image, PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    def __init__(self, db=None):
        self.db = db
        self.styles = getSampleStyleSheet()

    def to_csv(self, data, headers, filepath):
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in data:
                writer.writerow([row.get(h, "") for h in headers])

    def to_csv_bytes(self, data, headers):
        import io
        output = BytesIO()
        wrapper = io.TextIOWrapper(output, encoding="utf-8-sig")
        writer = csv.writer(wrapper)
        writer.writerow(headers)
        for row in data:
            writer.writerow([str(row.get(h, "")) for h in headers])
        wrapper.detach()
        output.seek(0)
        return output

    def to_excel(self, data, headers, filepath, title="Report"):
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="DADCE0"),
            right=Side(style="thin", color="DADCE0"),
            top=Side(style="thin", color="DADCE0"),
            bottom=Side(style="thin", color="DADCE0"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                if isinstance(value, float):
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )

        for col_idx, header in enumerate(headers, 1):
            column_width = max(
                len(str(header)),
                max(
                    (len(str(row.get(header, ""))) for row in data),
                    default=0,
                ),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(column_width + 4, 40)

        wb.save(filepath)

    def to_excel_bytes(self, data, headers, title="Report"):
        output = BytesIO()
        self.to_excel_bytes_io(output, data, headers, title)
        return output

    def to_excel_bytes_io(self, output, data, headers, title="Report"):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="DADCE0"),
            right=Side(style="thin", color="DADCE0"),
            top=Side(style="thin", color="DADCE0"),
            bottom=Side(style="thin", color="DADCE0"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                if isinstance(value, float):
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )

        wb.save(output)

    def to_pdf(self, data, headers, filepath, title="Report", subtitle=""):
        doc = SimpleDocTemplate(
            filepath, pagesize=landscape(letter),
            rightMargin=30, leftMargin=30,
            topMargin=30, bottomMargin=30,
        )
        elements = []

        title_style = ParagraphStyle(
            "CustomTitle", parent=self.styles["Title"],
            fontSize=18, spaceAfter=6,
            textColor=colors.HexColor("#1A73E8"),
        )
        elements.append(Paragraph(title, title_style))

        if subtitle:
            sub_style = ParagraphStyle(
                "CustomSub", parent=self.styles["Normal"],
                fontSize=10, textColor=colors.HexColor("#888888"),
                spaceAfter=12,
            )
            elements.append(Paragraph(subtitle, sub_style))

        date_style = ParagraphStyle(
            "DateStyle", parent=self.styles["Normal"],
            fontSize=8, textColor=colors.HexColor("#AAAAAA"),
            spaceAfter=16,
        )
        elements.append(
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style)
        )

        elements.append(Spacer(1, 0.1 * inch))

        table_data = [[self._clean_text(h) for h in headers]]
        for row in data:
            table_data.append([self._format_value(row.get(h, "")) for h in headers])

        col_widths = [
            max(len(str(h)) * 7 + 20, 60) for h in headers
        ]
        total_width = sum(col_widths)
        page_width = landscape(letter)[0] - 60
        if total_width > page_width:
            col_widths = [int(w * page_width / total_width) for w in col_widths]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A73E8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DADCE0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

        if len(table_data) > 35:
            elements.append(PageBreak())

        doc.build(elements)

    def to_pdf_bytes(self, data, headers, title="Report", subtitle=""):
        output = BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=landscape(letter),
            rightMargin=30, leftMargin=30,
            topMargin=30, bottomMargin=30,
        )
        elements = []

        title_style = ParagraphStyle(
            "CustomTitle", parent=self.styles["Title"],
            fontSize=18, spaceAfter=6,
            textColor=colors.HexColor("#1A73E8"),
        )
        elements.append(Paragraph(title, title_style))

        if subtitle:
            sub_style = ParagraphStyle(
                "CustomSub", parent=self.styles["Normal"],
                fontSize=10, textColor=colors.HexColor("#888888"),
                spaceAfter=12,
            )
            elements.append(Paragraph(subtitle, sub_style))

        date_style = ParagraphStyle(
            "DateStyle", parent=self.styles["Normal"],
            fontSize=8, textColor=colors.HexColor("#AAAAAA"),
            spaceAfter=16,
        )
        elements.append(
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style)
        )

        elements.append(Spacer(1, 0.1 * inch))

        table_data = [[self._clean_text(h) for h in headers]]
        for row in data:
            table_data.append([self._format_value(row.get(h, "")) for h in headers])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A73E8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DADCE0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

        doc.build(elements)
        output.seek(0)
        return output

    def _clean_text(self, text):
        text = str(text)
        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        text = text.replace('"', "&quot;").replace("'", "&#x27;")
        return text

    def _format_value(self, value):
        if value is None:
            return ""
        if isinstance(value, float):
            return f"{value:.2f}"
        return str(value)
